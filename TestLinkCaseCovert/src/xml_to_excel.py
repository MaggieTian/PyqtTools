import os,shutil
import re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
# 定义所需数据所在的列
from openpyxl.styles import Alignment,Font

ROW = 17
CASE_NUM = 2
CASE_COL = CASE_NUM+1             # Case开始的列
EXPECTED_RESULT_COL = CASE_COL+1  # 期望结果
PRIORITY_COL = CASE_COL + 5       # 优先级
HTML_CHAR = ["<div>", "</div>", "<p>", "</p>", "&nbsp;"]  # HTML中的字符（标签）
ACTION_STEPS = " 测试步骤：" + "\n"
'''
translate xml file to excel file 
'''


class XmlToExcel():
    def __init__(self, xml_path):
        self.sheet = ""
        self.xml_path = xml_path

    # 生成excel每行的内容
    def generate_excel(self, xml_path):
        '''

        :param xml_path: the path of xml file
        :return:  a generator that can generate excel case in every teset case
        '''
        if not xml_path:
            raise Exception("generate_excel函数中传入路径不能为空或者None")
        # 开始解析xml文件
        tree = ET.ElementTree(file=xml_path)
        root = tree.getroot()                                       # 获取根节点
        testcase_id = 0
        self.sheet = root.attrib['name']                            # excel sheet名为xml中testsuite name
        for child_root in root:
            if child_root.tag == "testcase":
                if "internalid" in child_root.attrib:
                    testcase_id = child_root.attrib['internalid']    # 写入testcase id
                else:                                                # 没有internalid时，就自动计数
                    testcase_id += 1

                testcase_title = ''
                if "name" in child_root.attrib:
                    testcase_title = child_root.attrib['name']       # 测试用例标题
                steps = ''                                           # 测试步骤
                expect_result = ''                                   # 期望结果
                priority = ''                                        # 优先级
                # 查找优先级
                for i in child_root.iterfind("importance"):
                    if i.text and int(i.text) == 3:
                        priority = 'H'

                # 利用xpath查找测试步骤和期望结果
                for i in child_root.iterfind("steps/step"):
                    for step in i.iterfind("actions"):
                        if step.text:
                            steps += step.text
                    for result in i.iterfind("expectedresults"):
                        if result.text:
                            expect_result += result.text.replace("\n", '')
                # 利用段落标签分割出测试步骤
                steps = self.remove_space_and_wrap(self.remove_html_char(steps))    # 去掉空格和html字符
                step_item = re.split(r"\d\.|\d、", steps)                          # 测试步骤以数字.开始或者数字、开始，eg:1. or 1、
                step_content = ''
                index = 1                       # 步骤计数
                for step in step_item:
                    if self.remove_space_and_wrap(step):
                        step_content += str(index)+"." + step + "\n"
                        index += 1
                steps = self.remove_space_and_wrap(testcase_title)+'\n' + ACTION_STEPS + step_content
                expect_result = self.remove_html_char(expect_result + '\n')
                yield {"tesetcase_id": testcase_id, "tesetcase": steps, "expected_result": expect_result, "priority": priority}
        yield {"sheet_name": self.sheet}

    # 去除xml文件中存在的一些html标签
    def remove_html_char(self, line):
        if line:
            result = line
            for char in HTML_CHAR:
                result = result.replace(char, '')
            return result
        else:
            return ""

    # 去除空格换行符
    def remove_space_and_wrap(self, line):
        if line:
            return line.replace("\t", "").replace("\n", "").replace(" ", '')
        else:
            return ""

    # 设置单元格style
    def set_style(self, row, style):
        for key, value in style.items():
            if key == "alignment":
                row.alignment = value
            if key == "font":
                row.font = value
            if key == "fill":
                row.fill = value
            if key == "border":
                row.border = value
            if key == "protection":
                row.protection = value

    def write_to_excel(self, export_dir, content, name):
        '''
        :param export_dir: the directory that save the result excel file
        :param content:  the xml content after converting
        :param name: the result excel file name
        :return: None
        '''

        # 得到存放excel模板路径
        # 将模板中的内容复制到新建的excel文件中
        shutil.copyfile(os.path.abspath("template.xlsx"), os.path.join(export_dir, name + ".xlsx"))
        workbook = load_workbook(os.path.join(export_dir, name + ".xlsx"))          # 打开文件
        sheet = workbook.active                                                     # 获取正在显示的sheet
        row = ROW                                                                   # tesetcase开始的行
        alignment = Alignment(horizontal='general', vertical='bottom', text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0)  # 设置单元格中文字自动换行
        font = Font(name="宋体", size=10)   # 设置字体和大小
        style = {"alignment": alignment, "font": font}
        # 将用例写进excel
        for line in content:
            if "sheet_name" not in line.keys():
                # 写入ID
                cell = sheet.cell(row, CASE_NUM)
                cell.value = int(line['tesetcase_id'])
                self.set_style(cell, style)
                # 写入测试标题和步骤
                cell = sheet.cell(row, CASE_COL)
                cell.value = line['tesetcase']
                self.set_style(cell, style)
                # 写入期望结果
                cell = sheet.cell(row, EXPECTED_RESULT_COL)
                cell.value = line['expected_result']
                self.set_style(cell, style)
                # 写入优先级
                cell = sheet.cell(row, PRIORITY_COL)
                cell.value = line['priority']
                self.set_style(cell, style)
                row += 1
            else:
                sheet.title = line["sheet_name"]                    # 修改sheet名字为xml中test suite名
        workbook.save(os.path.join(export_dir, name + ".xlsx"))     # 保存文件


if __name__ == "__main__":
    p = XmlToExcel()
    text = p.generate_excel(r"C:\Users\qtian\Desktop\autotest\PyqtTools\TestLinkCaseCovert\template\Read-only file system.xml")
    p.write_to_excel(r"C:\Users\qtian\Desktop\autotest\PyqtTools\TestLinkCaseCovert\template", text, "test_result")